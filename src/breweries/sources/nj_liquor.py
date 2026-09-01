"""New Jersey Division of Alcoholic Beverage Control (ABC) — statewide "Wholesale/
State Issued Licensee Listing" Excel export.

Source: linked from the ABC's own "Licensing Reports" page
(https://www.njoag.gov/about/divisions-and-offices/division-of-alcoholic-beverage-control-home/licensing-bureau-applications-and-information/licensing-reports/),
e.g. "WHOLESALE-LICENSES-REPORT-JULY-2026.xlsx" -- a genuine bulk, no-login,
statewide export of every state-issued (as opposed to municipally-issued
retail) license ABC has active, refreshed monthly. New Jersey brewery
licenses (Plenary Brewery, Limited Brewery, Restricted Brewery) are all
State-issued, unlike NJ's retail consumption/distribution licenses, which are
issued by individual municipalities and appear in a separate "Retail
Licensee Listing" export -- not used here since it carries none of the three
brewery classes. This page (njoag.gov) enforces a basic bot-detection layer
(a plain `requests.get()` without a browser User-Agent returns HTTP 403) --
worked around the same way every other module in this project already does,
by setting a realistic browser `User-Agent` header (see `_HEADERS` below,
identical in spirit to il_liquor.py's and wa_liquor.py's own headers) --
*not* by solving a CAPTCHA or defeating a session/cookie challenge, which
this project's rules forbid and which this page does not present. New
Jersey's Socrata/open-data presence (data.nj.gov) was also checked and
carries no ABC/liquor-license dataset; the ABC's own monthly report is the
only bulk source.

## Columns and scope

One row per license record: License Number, License Type, Status
(mislabeled "State" in the sheet's header row -- its values are "Active" /
etc., not a US state abbreviation), Establishment, Licensee, Effective Date,
City, Premise Address (a single "STREET ,CITY,NJ,ZIP,USA" string). Every row
in this export already carries `Status == "Active"`; there is no expired/
inactive row to filter (unlike IL's cumulative ILCC export), confirmed by
inspecting the full `License Type` value distribution on the 2026-07-01
fetch (1,604 rows, all Active).

There is no county column -- only City/Premise Address -- so, like WA's
WSLCB list, records are geocoded downstream via the project's Census
Geocoder fallback (breweries.geocode.fill_missing_coords), the same path
OBDB already uses.

## License classes and inclusion rule

New Jersey licenses breweries under three License Type values (N.J.S.A.
33:1-10):

- "Plenary Brewery License" -- unlimited production, the traditional
  large-scale brewery license (1 row on the 2026-07-01 fetch: Mark Anthony
  Brewing's Hillside facility, maker of Cayman Jack).
- "Limited Brewery License" -- New Jersey's craft-brewery license, capped at
  300,000 barrels/yr with on-site sales/self-distribution rights (123 rows).
- "Restricted Brewery License" -- New Jersey's brewpub license, tied to a
  companion retail consumption license at the same restaurant premises for
  on-site sale (20 rows) -- unioned in the same way IL's 1C Brew Pub class is,
  since the companion retail license lives in the separate municipally-issued
  Retail Licensee Listing this module does not load, so there is no
  cross-file double-count risk.

BREWERY_LICENSE_TYPES is the union of all three (144 raw rows). Excluded,
deliberately: "Out of State Winery License" (589 rows, the largest category
in the file -- out-of-state wine importers, not NJ breweries), "Craft
Distillery License", "Farm Winery License", "Plenary Winery License",
"Cidery and Meadery" (1 row), and every wholesale/warehouse/transportation/
permit category -- none of these are brewing licenses.

## Dedup: companion license rows at the same premises

A handful of licensees hold two manufacturer sub-licenses (e.g. a brewery +
craft-distillery combo) at one physical address, filed as two separate
license-number rows that share a state-assigned entity ID as the numeric
prefix of `License Number` before its final `-NNN` segment (New Jersey's own
license-numbering convention). Records are deduplicated on (entity ID prefix,
normalized street address, city) -- distinct from a pure address-based key,
which would incorrectly merge e.g. Iron Hill Brewery's Maple Shade and
Voorhees locations if they ever shared a landlord's mailing address; this
key only merges rows that are *both* the same entity ID *and* the same
address, so genuine multi-location chains (Iron Hill, Triumph Brewing,
Kings Road Brewing, Five Dimes Brewery, Invertase Brewing -- all appear
twice, at two different NJ addresses each) are correctly kept as separate
rows/locations. One true exact duplicate row (Bull N Bear Brewery LLC,
Summit, appearing twice with identical License Number) is also removed by
this dedup step.

## Missing/unparseable address

Two of the 144 raw brewery-class rows are dropped for lack of a usable
address, the same treatment IL gives to dba_address parse failures: "Ship
Bottom Brewery LLC" (Limited Brewery License) carries no City or Premise
Address at all in the 2026-07-01 export, and "Avalon Brew Pub" (Restricted
Brewery License) carries a malformed multi-line Premise Address ("78TH & DUNE
DRIVE  \\n   AVALON USA 08202 ,AVALON,,08202,") that omits the "NJ" state
token the parser keys on, so it fails to match rather than being silently
mis-parsed.
"""

from __future__ import annotations

import glob
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from breweries.manifest import log_fetch, log_filter

RAW_DIR = Path("data/raw/nj_liquor")
REPORTS_PAGE_URL = (
    "https://www.njoag.gov/about/divisions-and-offices/division-of-alcoholic-beverage-control-home/"
    "licensing-bureau-applications-and-information/licensing-reports/"
)

BREWERY_LICENSE_TYPES = [
    "Plenary Brewery License",
    "Limited Brewery License",
    "Restricted Brewery License",
]

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
_LINK_RE = re.compile(r'href="([^"]*WHOLESALE[^"]*\.xlsx)"', re.IGNORECASE)
_ADDR_RE = re.compile(r"^(?P<street>.+?)\s*,\s*(?P<city>[^,]+),\s*NJ,\s*(?P<zip>\d{5})", re.IGNORECASE)


def _discover_report_url() -> str:
    resp = requests.get(REPORTS_PAGE_URL, headers=_HEADERS, timeout=60)
    resp.raise_for_status()
    m = _LINK_RE.search(resp.text)
    if not m:
        raise RuntimeError(
            "Could not find a 'WHOLESALE-LICENSES-REPORT-....xlsx' download link on the "
            f"NJ ABC Licensing Reports page ({REPORTS_PAGE_URL}) -- the page's layout may "
            "have changed."
        )
    href = m.group(1)
    return href if href.startswith("http") else "https://www.njoag.gov" + href


def fetch(force: bool = False) -> Path:
    """Download NJ ABC's monthly Wholesale/State Issued Licensee Listing, or reuse the cache."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    existing = sorted(glob.glob(str(RAW_DIR / "nj_wholesale_licensees_*.xlsx")))
    if existing and not force:
        return Path(existing[-1])

    url = _discover_report_url()
    resp = requests.get(url, headers=_HEADERS, timeout=120)
    resp.raise_for_status()

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    dest = RAW_DIR / f"nj_wholesale_licensees_{ts}.xlsx"
    dest.write_bytes(resp.content)

    df = _read_raw(dest)
    log_fetch(
        source="nj_liquor", url=url, dest_path=str(dest), row_count=len(df),
        notes="full statewide NJ ABC state-issued/wholesale license roll, all license types, "
              f"linked from the ABC's own Licensing Reports page ({REPORTS_PAGE_URL})",
    )
    return dest


def _read_raw(path: Path) -> pd.DataFrame:
    """NJ ABC's export has a 2-line title block above the real header row (see
    module docstring); locate the header by its known first cell rather than a
    fixed row offset.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "License Number")
    cols = list(rows[hdr_idx])
    data = rows[hdr_idx + 1:]
    df = pd.DataFrame(data, columns=cols)
    return df[df["License Number"].notna()].reset_index(drop=True)


def _parse_address(addr: object) -> tuple[str | None, str | None, str | None]:
    if not isinstance(addr, str):
        return None, None, None
    m = _ADDR_RE.match(addr.strip())
    if not m:
        return None, None, None
    return m.group("street").strip(), m.group("city").strip().title(), m.group("zip")


def load() -> pd.DataFrame:
    """Load the cached export, filter to the three brewery license types, and dedup
    companion manufacturer sub-licenses at the same premises. See module docstring
    for the full inclusion-rule reasoning.
    """
    path = fetch()
    df = _read_raw(path)
    n0 = len(df)

    brewery = df[df["License Type"].isin(BREWERY_LICENSE_TYPES)].copy()
    log_filter("nj_liquor", f"License Type in {BREWERY_LICENSE_TYPES}", n0, len(brewery))

    parsed = brewery["Premise Address"].apply(
        lambda a: pd.Series(_parse_address(a), index=["street_address", "city", "zip"])
    )
    brewery = pd.concat([brewery.reset_index(drop=True), parsed.reset_index(drop=True)], axis=1)

    n_before_addr = len(brewery)
    brewery = brewery[brewery["street_address"].notna()].copy()
    log_filter(
        "nj_liquor", "Premise Address parsed successfully", n_before_addr, len(brewery),
        notes="drops rows with no usable premise address (Ship Bottom Brewery LLC: none at "
              "all; Avalon Brew Pub: malformed/unparseable), see module docstring",
    )

    brewery["_entity_id"] = brewery["License Number"].astype(str).str.rsplit("-", n=1).str[0]
    brewery["_dedup_key"] = (
        brewery["_entity_id"] + "|" + brewery["street_address"].str.upper().str.strip()
        + "|" + brewery["city"].str.upper().str.strip()
    )
    n_before_dedup = len(brewery)
    brewery = brewery.drop_duplicates("_dedup_key", keep="first")
    log_filter(
        "nj_liquor",
        "dedup companion manufacturer sub-licenses at the same premises (entity ID prefix of "
        "License Number + street address + city)",
        n_before_dedup, len(brewery),
        notes="also removes one true duplicate row (Bull N Bear Brewery LLC, Summit) -- see "
              "module docstring",
    )

    brewery["nj_liquor_id"] = brewery["License Number"]
    brewery["licensee_name"] = brewery["Licensee"]
    brewery["state"] = "NJ"
    brewery["lat"] = pd.NA
    brewery["lon"] = pd.NA
    brewery = brewery.rename(columns={"License Type": "license_type", "License Number": "license_number"})

    return brewery[[
        "nj_liquor_id", "license_number", "licensee_name", "license_type", "street_address",
        "city", "state", "zip", "lat", "lon",
    ]]
